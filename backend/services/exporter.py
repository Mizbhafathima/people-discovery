import io
import json
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

COLUMNS = [
    ("Name", "name"),
    ("Email", "email"),
    ("Phone", "phone"),
    ("Job Title", "job_title"),
    ("LinkedIn URL", "linkedin_url"),
    ("Instagram URL", "instagram_url"),
    ("Twitter URL", "twitter_url"),
    ("Source URL", "source_url"),
    ("Confidence", "confidence"),
    ("Domain", "domain"),
]


class ExporterService:
    def people_to_dicts(self, people_orm_list) -> List[dict]:
        results = []
        for person in people_orm_list:
            row = {}
            for _, key in COLUMNS:
                row[key] = getattr(person, key, None)
            results.append(row)
        return results

    def to_json(self, people: List[dict]) -> str:
        return json.dumps(people, indent=2, ensure_ascii=False)

    def to_excel_bytes(self, people: List[dict]) -> bytes:
        workbook = Workbook()
        ws = workbook.active
        ws.title = "People"

        ws.cell(row=1, column=1, value="All Domains")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
        title_cell = ws.cell(row=1, column=1)
        title_cell.font = Font(bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(fill_type="solid", fgColor="1F3864")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, (header, _) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(fill_type="solid", fgColor="D9D9D9")

        for row_idx, person in enumerate(people, start=3):
            fill_color = "FFFFFF" if row_idx % 2 == 0 else "EBF3FF"
            for col_idx, (_, key) in enumerate(COLUMNS, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=person.get(key))
                cell.fill = PatternFill(fill_type="solid", fgColor=fill_color)

        for col_idx, (_, key) in enumerate(COLUMNS, start=1):
            max_length = len(str(ws.cell(row=2, column=col_idx).value or ""))
            for person in people:
                value = person.get(key)
                max_length = max(max_length, len(str(value)) if value is not None else 0)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(60, max(12, max_length + 4))

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()
